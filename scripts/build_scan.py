# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD = ["New Label","Study (1st author, year)","Site / Lab","Scanner (manufacturer / model)",
        "Field Strength","Functional sequence","TR (ms)","TE (ms)","Flip angle (deg)",
        "Voxel size (acquired)","Slices","FOV","Smoothing FWHM","Status","Source / Notes"]

R=[]
def add(*v): R.append(list(v))
NR="NOT RETRIEVED"

# ---- verified ----
add("Atlas_2010_expectancy","Atlas 2010","Columbia (Wager)","GE Signa TwinSpeed Excite HD","1.5 T",
"Spiral in/out gradient-echo","2000","40","84","3.5 x 3.5 x 4.55 mm (interp 2 mm)","24","224 mm; 64x64","8 mm","Verified",
"J Neurosci 2010. PMC2966558")

add("Atlas_2012_remifentanil","Atlas 2012","Columbia (Wager)","","","","","","","","","","",NR,
"J Neurosci 2012; PMC3387557 - access blocked (reCAPTCHA). Verify from full text.")

add("Brascher_2016_control","Brascher 2016","McGill / MNI (Schweinhardt)","","","","","","","","","","",NR,
"J Neurosci 2016; PMC6601855 - access blocked. Likely 3T Siemens Trio (same site as Becker 2017) - verify.")

add("Becker_2017_reward","Becker 2017","McGill / MNI","Siemens Trio (32-ch)","3 T","T2*-weighted EPI (whole brain)",
"2370","30","90","3 x 3 x 3 mm (45 slices, descending)","45","192 x 192 mm; 64x64","(not stated in excerpt)","Verified",
"SCAN 2017. PMC5390724")

add("Bingel_2011_remifentanil_posExpectancy","Bingel 2011","Hamburg/Oxford (Bingel/Tracey)","","","","","","","","","","",NR,
"Sci Transl Med 2011 - paywalled. Verify (Oxford 3T Siemens Trio likely).")

add("Woo_2015_selfregulation","Woo 2015","Columbia PICS (Wager)","Philips Achieva TX","3 T","EPI (parallel imaging, SENSE 1.5)",
"2000","20","(not stated)","3 x 3 x 3 mm (interp 2 mm)","42 (interleaved)","224 mm; 64x64","8 mm","Verified",
"PLoS Biol 2015. PMC4285399")

add("Habermann_2025_control / _predictability","Habermann 2025","Hamburg (Buchel)","Siemens Prisma (64-ch)","3 T",
"EPI (multiband x2, GRAPPA x2)","1500","26","60","~2 mm (FOV 224; voxel n/r in excerpt)","(n/r)","224 mm","(n/r)","Verified",
"Nat Commun 2025. PMC12627469. One acquisition for both control & predict contrasts.")

add("Jepma_2015_conceptualConditioning","Jepma & Wager 2015","Boulder (Wager)","","","","","","","","","","",NR,
"Psychol Sci 2015 - paywalled. NOTE: may be behavioural/SCR (confirm whether fMRI dataset).")

add("Woo_2017_expectancy / _perceivedControl","Woo 2017 (SIIPS-1)","Multiple (CANlab pooled)","Multiple scanners","Mixed",
"Mixed (pooled datasets)","-","-","-","-","-","-","-","Multi-study",
"Nat Commun 2017; ncomms14211. SIIPS-1 pools 6 datasets (Studies 1-6) - no single protocol. Verify which dataset 'ilcp' maps to.")

add("Jepma_2018_expectancy","Jepma 2018","Boulder (Wager)","","","","","","","","","","",NR,
"Nat Hum Behav 2018; PMC6768437 - access blocked. Verify (Boulder 3T).")

add("Koban_2019_socialExpectancy","Koban 2019","Boulder (Wager)","Siemens TrioTim","3 T","T2* EPI (GRAPPA)",
"1300","25","50","3.4 mm isotropic (resampled 3 mm)","26 (interleaved)","220 mm","8 mm","Verified",
"Nat Commun 2019. PMC6736972")

add("Kober_2019_mindfulAcceptance","Kober 2019","Columbia (Ochsner/Wager)","GE Signa TwinSpeed Excite HD","1.5 T","T2*-weighted EPI BOLD",
"2000","34","90","3.5 mm in-plane x 4.5 mm","28","224 mm; 64x64","(via SPGR; n/r)","Verified",
"SCAN 2019. PMC7057281")

add("LopezSola_2018_prosocialMeaning","Lopez-Sola 2018","Boulder (Wager)","","","","","","","","","","",NR,
"Psychosom Med 2018; PMC6218300 - access blocked. Verify.")

add("LopezSola_2019_handholding","Lopez-Sola 2019","Boulder (Wager)","","","","","","","","","","",NR,
"Pain 2019 - paywalled. Verify (Boulder 3T Siemens).")

add("BotvinikNezer_2024_placebo","Botvinik-Nezer 2024","CU Boulder (Wager)","Siemens Prisma (32-ch)","3 T","EPI (multiband x8)",
"460","27.20","44","2.7 x 2.7 x 2.7 mm","56","220 mm","(n/r)","Verified",
"Nat Commun 2024. PMC11255344")

add("Roy_2009_emotionModulation","Roy 2009","Montreal (Rainville)","","","","","","","","","","",NR,
"PNAS 2009; PMC2779826 - access blocked. Combined brain + spinal cord fMRI (electrical sural-nerve pain). Verify.")

add("Steininger_2025_natureExposure","Steininger 2025","Vienna (Lamm)","Siemens Magnetom Skyra (32-ch)","3 T","EPI (multiband x4)",
"800","34","50","2.18 mm (matrix 96x96x36)","36","210 x 210 x 138 mm","(n/r)","Verified",
"Nat Commun 2025. doi:10.1038/s41467-025-56870-2. PMC11906725")

add("Kim_2024_cueIntegration","Kim 2024","Sungkyunkwan (Woo)","Siemens Magnetom Prisma (64-ch)","3 T","T2* EPI (multiband x8)",
"460","27.20","44","2.7 x 2.7 x 2.7 mm","56","220 mm","5 mm","Verified",
"Sci Adv 2024. doi:10.1126/sciadv.ado8230. PMC11389792. T1 MPRAGE: TR2400/TE2.34/flip8/0.7mm. NOTE: original DOI adk7421 was incorrect; corrected to ado8230.")

add("Tinnermann_2022_remifentanil","Tinnermann 2022","Hamburg (Buchel)","Siemens TIM Trio","3 T","Combined brain+spinal-cord EPI",
"2650 (brain)","30 brain / 34 cord","(n/r)","Brain 2.0 mm iso; cord 1.2 x 1.2 x 3.5 mm","32 brain / 10 cord","220 brain / 132 cord","6 mm (brain)","Verified",
"eLife 2022;11:e74293")

add("Riegner_2023_mindfulness","Riegner 2023","UCSD (Zeidan)","","","","","","","","","","",NR,
"Pain 2023; PMC9823141 - access blocked. Verify (UCSD 3T Siemens).")

# ---- placebo (Zunhammer 2018 datasets) ----
add("Bingel_2006_placebo","Bingel 2006","Hamburg (Buchel)","","","","","","","","","","",NR,
"Pain 2006 - paywalled. Laser pain. Verify.")
add("Bingel_2011_placebo","Bingel 2011","Hamburg/Oxford","","","","","","","","","","",NR,
"Sci Transl Med 2011 - paywalled (same study as remi row). Verify.")
add("Choi_2011_placebo","Choi 2011","Korea (Choi)","","","","","","","","","","",NR,
"Choi 2011 - paywalled. Electrical pain. Verify citation + params.")
add("Eippert_2009_placebo","Eippert 2009","Hamburg (Buchel)","Siemens Trio (8-ch)","3 T","T2*-weighted EPI",
"2230","25","80","3 x 3 mm in-plane; 2 mm slice + 1 mm gap","38","192 x 192 mm; 64x64","8 mm","Verified",
"Neuron 2009. PMC6670627")
add("Ellingsen_2013_placebo","Ellingsen 2013","Oslo / Boston","Philips Achieva","3 T","(EPI; details in SI)",
"","","","","","","",NR + " (partial)",
"PNAS 2013. PMC3816412. Scanner+field verified; TR/TE/voxel in Supplementary Info - verify.")
add("Elsenbruch_2012_placebo","Elsenbruch/Kotsis 2012","Essen (Elsenbruch)","","","","","","","","","","",NR,
"NeuroImage 2012 - paywalled. Visceral (rectal distension). Verify.")
add("Freeman_2015_placebo","Freeman 2015","MGH (Kong/Gollub)","","","","","","","","","","",NR,
"NeuroImage 2015; PMC4408248 - access blocked. Verify.")
add("Geuter_2013_placebo","Geuter 2013","Hamburg (Buchel)","","","","","","","","","","",NR,
"NeuroImage 2013; PMC3578963 - access blocked. Likely 3T Siemens Trio - verify.")
add("Kong_2006_placebo","Kong 2006","MGH (Gollub/Kong)","Siemens (head-only)","3 T","EPI",
"2000","40","90","3.13 x 3.13 mm; 4 mm slice + 1 mm skip","30","(n/r); AC-PC","8 mm","Verified",
"J Neurosci 2006. PMC6674420")
add("Kong_2009_placebo","Kong 2009","MGH (Kong)","","","","","","","","","","",NR,
"NeuroImage 2009; PMC2737445 - access blocked. Sham acupuncture. Verify.")
add("Lui_2010_placebo","Lui 2010","Modena (Porro)","","","","","","","","","","",NR,
"Pain 2010 - paywalled. Laser pain, sham TENS. Verify.")
add("Schenk_2014_placebo","Schenk 2014","Hamburg (Buchel)","","","","","","","","","","",NR,
"Pain 2014 - paywalled. Heat + capsaicin. Likely 3T Siemens Trio - verify.")
add("Theysohn_2014_placebo","Theysohn (Essen)","Essen (Elsenbruch)","","","","","","","","","","",NR,
"Visceral placebo dataset - paywalled. Verify exact citation + params.")
add("Wager_2004_study1_placebo","Wager 2004 (Study 1)","Michigan/Princeton (Wager)","","","","","","","","","","",NR,
"Science 2004 - paywalled, no PMC. Electrical shock. Verify.")
add("Wager_2004_study2_placebo","Wager 2004 (Study 2)","Michigan/Princeton (Wager)","","","","","","","","","","",NR,
"Science 2004 - paywalled, no PMC. Heat. Verify.")
add("Wrobel_2014_placebo","Wrobel 2014","Hamburg (Bingel)","","","","","","","","","","",NR,
"Cortex 2014 - paywalled. Verify (3T Siemens).")

wb=Workbook(); ws=wb.active; ws.title="Scanning parameters"
hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(name="Arial",bold=True,color="FFFFFF",size=10)
cf=Font(name="Arial",size=10); thin=Side(style="thin",color="D9D9D9")
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ver=PatternFill("solid",fgColor="E2EFDA"); nrf=PatternFill("solid",fgColor="FCE4D6"); multif=PatternFill("solid",fgColor="FFF2CC")
for j,h in enumerate(HEAD,1):
    c=ws.cell(1,j,h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bd
for i,row in enumerate(R,2):
    for j,val in enumerate(row,1):
        c=ws.cell(i,j,val); c.font=cf; c.border=bd; c.alignment=Alignment(vertical="top",wrap_text=True)
    st=row[13]
    fill = ver if st=="Verified" else (multif if st=="Multi-study" else nrf)
    ws.cell(i,14).fill=fill
widths=[30,20,22,26,12,24,10,12,12,26,14,16,14,16,40]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes="B2"; ws.row_dimensions[1].height=40
ws.auto_filter.ref=f"A1:{get_column_letter(len(HEAD))}{len(R)+1}"

# summary sheet
sm=wb.create_sheet("README")
ver_n=sum(1 for r in R if r[13]=="Verified")
lines=[
 ["Scanning parameters - companion to Study_summary_completed.xlsx",""],
 ["",""],
 ["Acquisition parameters extracted from each paper's Methods section.",""],
 ["Verified rows = transcribed from the actual full-text Methods.",""],
 ["NOT RETRIEVED = full text paywalled or temporarily access-blocked (PMC reCAPTCHA).",""],
 ["For blocked rows the Site/Lab is given to ease look-up; do NOT assume a scanner from site alone.",""],
 ["",""],
 ["Verified papers", str(ver_n)],
 ["Not retrieved / partial", str(sum(1 for r in R if r[13].startswith('NOT')))],
 ["Multi-study (no single protocol)", str(sum(1 for r in R if r[13]=='Multi-study'))],
 ["",""],
 ["Field-strength snapshot (verified rows):",""],
 ["1.5 T","Atlas 2010, Kober 2019"],
 ["3 T","Becker 2017, Woo 2015, Habermann 2025, Koban 2019, Botvinik-Nezer 2024, Steininger 2025, Kim 2024, Tinnermann 2022, Eippert 2009, Kong 2006"],
 ["",""],
 ["'(n/r)' = not reported in the text excerpt retrieved (may be in supplement).",""],
 ["Next step: blocked rows can be retried later (reCAPTCHA resets) or filled via institutional access.",""],
]
for i,row in enumerate(lines,1):
    for j,val in enumerate(row,1):
        c=sm.cell(i,j,val); c.font=Font(name="Arial",size=10,bold=(i==1)); c.alignment=Alignment(vertical="top",wrap_text=True)
sm.column_dimensions["A"].width=44; sm.column_dimensions["B"].width=80
sm.cell(1,1).font=Font(name="Arial",size=13,bold=True)

wb.save("Scanning_parameters.xlsx")
print("Saved Scanning_parameters.xlsx | rows:",len(R),"| verified:",ver_n)
