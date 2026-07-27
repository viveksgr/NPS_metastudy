# -*- coding: utf-8 -*-
# consolidate.py
#  1. Repair column-shift corruption in 'Studies' (N landed in Population; notes in orphan col18)
#  2. Remove redundant columns (orphan col18 in Studies; 'Study (1st author, year)' in Scanning)
#  3. Add 2 moderator columns to 'Studies': Stimulus class, Stimulus calibration
import io, sys, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN = "Study_summary_completed.xlsx"

HFILL = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CFONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

POP_KW = ("healthy","adult","women","woman","men","man","male","female","naive",
          "participant","volunteer","child","patient")

# Calibration by New Label (fixed vs individually calibrated stimulus)
CALIB = {
 "Atlas_2010_expectancy":"Calibrated","Atlas_2012_remifentanil":"Calibrated",
 "Becker_2017_reward":"Calibrated","Bingel_2011_remifentanil_posExpectancy":"Calibrated",
 "Woo_2015_selfregulation":"Fixed","Habermann_2025_control":"Calibrated",
 "Habermann_2025_predictability":"Calibrated","Jepma_2018_expectancy_study1":"Calibrated",
 "Woo_2017_expectancy":"Calibrated","Woo_2017_perceivedControl":"Calibrated",
 "Jepma_2018_expectancy_study2":"Calibrated","Koban_2019_socialExpectancy":"Calibrated",
 "Kober_2019_mindfulAcceptance":"Calibrated","LopezSola_2018_prosocialMeaning":"Calibrated",
 "LopezSola_2019_handholding":"Calibrated","BotvinikNezer_2024_placebo":"Calibrated",
 "Roy_2009_emotionModulation":"Calibrated","Steininger_2025_natureExposure":"Calibrated",
 "Kim_2024_cueIntegration":"Calibrated","Tinnermann_2022_remifentanil":"Calibrated",
 "Riegner_2023_mindfulness":"Fixed","Bingel_2006_placebo":"Fixed? (verify)",
 "Bingel_2011_placebo":"Calibrated","Choi_2011_placebo":"(verify)",
 "Eippert_2009_placebo":"Calibrated","Ellingsen_2013_placebo":"Calibrated",
 "Elsenbruch_2012_placebo":"Calibrated","Freeman_2015_placebo":"Calibrated",
 "Geuter_2013_placebo":"Calibrated","Kong_2006_placebo":"Calibrated",
 "Kong_2009_placebo":"Calibrated","Lui_2010_placebo":"Fixed? (verify)",
 "Schenk_2014_placebo":"Calibrated","Theysohn_2014_placebo":"Calibrated",
 "Wager_2004_study1_placebo":"Calibrated","Wager_2004_study2_placebo":"Calibrated",
 "Wrobel_2014_placebo":"Calibrated","MPA2_heat_regulation":"(verify)",
 "MPA2_pressure_regulation":"(verify)",
}

def stim_class(modality):
    m = (modality or "").lower()
    if any(k in m for k in ("heat","laser","thermal","capsaicin")): return "Thermal"
    if any(k in m for k in ("electric","visceral","pressure","mechanical","distension")): return "Non-thermal"
    return "(verify)"

def is_population(v):
    if v is None: return True
    return any(k in str(v).lower() for k in POP_KW)

wb = openpyxl.load_workbook(FN)
ws = wb["Studies"]
maxr = ws.max_row

# ---------- PHASE 1: repair corruption ----------
n_fixN, n_fixNote = 0, 0
for r in range(2, maxr+1):
    pop = ws.cell(r,13).value          # current 'Population'
    if pop is not None and str(pop).strip() and not is_population(pop):
        # misplaced N -> move to col12 (N analyzed); restore Population
        ws.cell(r,12, pop)
        ws.cell(r,13, "Healthy adults")
        n_fixN += 1
    note18 = ws.cell(r,18).value       # orphan notes
    if note18 is not None and str(note18).strip():
        ws.cell(r,17, note18)          # move into real Notes
        n_fixNote += 1
print(f"Repaired N->col12 for {n_fixN} rows; moved notes->col17 for {n_fixNote} rows")

# delete orphan col18
ws.delete_cols(18)

# ---------- PHASE 2: add 2 moderator columns ----------
c_class = ws.max_column + 1     # = 18
c_cal   = c_class + 1           # = 19
for col, title in [(c_class,"Stimulus class (thermal / non-thermal)"),
                   (c_cal,  "Stimulus calibration (fixed / calibrated)")]:
    c = ws.cell(1, col, title)
    c.fill = HFILL; c.font = HFONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BD
    ws.column_dimensions[get_column_letter(col)].width = 20

for r in range(2, maxr+1):
    label = ws.cell(r,3).value
    modality = ws.cell(r,9).value
    sc = stim_class(modality)
    cal = CALIB.get(label, "(verify)")
    for col,val in [(c_class,sc),(c_cal,cal)]:
        cc = ws.cell(r,col,val); cc.font=CFONT; cc.border=BD
        cc.alignment=Alignment(vertical="top", wrap_text=True)

# update autofilter
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{maxr}"
print(f"Added moderator cols: {c_class}=Stimulus class, {c_cal}=Calibration")

# ---------- PHASE 3: remove redundant col in Scanning ----------
sc_ws = wb["Scanning parameters"]
# delete col2 'Study (1st author, year)' (redundant with New Label)
hdr2 = sc_ws.cell(1,2).value
sc_ws.delete_cols(2)
sc_ws.auto_filter.ref = f"A1:{get_column_letter(sc_ws.max_column)}{sc_ws.max_row}"
print(f"Scanning: removed redundant column '{hdr2}'")

wb.save(FN)
print("Saved", FN)

# ---------- verify ----------
wb2 = openpyxl.load_workbook(FN, read_only=True)
s = wb2["Studies"]
print("\nStudies headers:")
for j,c in enumerate(next(s.iter_rows(max_row=1)),1):
    print(f"  col{j}: {c.value}")
print("\nSample (label | N | Population | StimClass | Calibration):")
for r in s.iter_rows(min_row=2, values_only=True):
    print(f"  {str(r[2])[:34]:35} | {str(r[11])[:22]:23} | {str(r[12])[:16]:17} | {str(r[17])[:11]:12} | {r[18]}")
