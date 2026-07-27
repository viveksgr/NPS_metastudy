# -*- coding: utf-8 -*-
# update_final.py - Choi 2011 (correct paper) + Woo 2017 Study 6 (from supplement)
import io, sys
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN = "Study_summary_completed.xlsx"
GREEN = PatternFill("solid", fgColor="E2EFDA")

# cols 3..15: Site, Scanner, Field, Sequence, TR, TE, Flip, Voxel, Slices, FOV, Smoothing, Status, Note
SCAN = {
"Choi_2011_placebo": ["Korea (Choi), Gwangju","ISOL Technology 3T (quadrature head coil)","3 T",
  "Gradient-echo EPI","3000","30","80","3.75 x 3.75 x 4 mm","35","240 x 240 mm; 64x64","5 mm","Verified",
  "NeuroReport 2011 'Placebo effects on analgesia related to testosterone and premotor activation' (Choi JC et al.). Correct paper now in folder."],

"Woo_2017_expectancy / _perceivedControl": ["CU Boulder - SIIPS-1 Study 6","Siemens Tim Trio","3 T",
  "EPI (interleaved, iPAT=2)","1980","25","75","3.4 x 3.4 x 3.0 mm","35 (interleaved)","220 mm; 64x64","8 mm","Verified",
  "SIIPS-1 Study 6 (N=29, 16F, age 20.4). Both ILCP rows (expectancy + perceivedControl) come from this single dataset. From Woo 2017 Suppl Table 3. iPAT=2, 5 discarded vols, SPM8. Resampled to 2 mm before smoothing."],
}

wb = openpyxl.load_workbook(FN)
ws = wb["Scanning parameters"]
labels = {ws.cell(r,1).value: r for r in range(2, ws.max_row+1)}
for label, vals in SCAN.items():
    r = labels.get(label)
    if not r:
        print("  !! not found:", label); continue
    for j, v in enumerate(vals, start=3):
        ws.cell(r, j, v)
    ws.cell(r,14).fill = GREEN
    print("  ok:", label)

# refresh README counts
if "README" in wb.sheetnames:
    rm = wb["README"]
    ver = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,14).value=="Verified")
    part = sum(1 for r in range(2, ws.max_row+1) if str(ws.cell(r,14).value).startswith(("Verified (partial)","NOT RETRIEVED")))
    for r in range(1, rm.max_row+1):
        a = rm.cell(r,1).value
        if a=="Verified papers": rm.cell(r,2,str(ver))
        elif a and a.startswith("Not retrieved"): rm.cell(r,2,str(part))

wb.save(FN)
print("Saved", FN)
cnt={}
for r in range(2, ws.max_row+1):
    s=ws.cell(r,14).value; cnt[s]=cnt.get(s,0)+1
print("\nFinal status counts:")
for k,v in sorted(cnt.items(), key=lambda x:-x[1]): print(f"  {v:3}  {k}")
