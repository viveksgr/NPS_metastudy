# -*- coding: utf-8 -*-
import io, sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN="Study_summary_completed.xlsx"
GREEN=PatternFill("solid",fgColor="E2EFDA"); ORANGE=PatternFill("solid",fgColor="FCE4D6"); GREY=PatternFill("solid",fgColor="D9D9D9")
HFILL=PatternFill("solid",fgColor="1F4E78"); HFONT=Font(name="Arial",bold=True,color="FFFFFF",size=10)
CFONT=Font(name="Arial",size=10); THIN=Side(style="thin",color="D9D9D9"); BD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

wb=openpyxl.load_workbook(FN); ws=wb["Studies"]
idx={ws.cell(r,3).value:r for r in range(2,ws.max_row+1)}

# ---- 1a. reassign 'ie' row -> Roy 2014 Study 3 ----
r=idx["Jepma_2018_expectancy_study1"]
vals={3:"Roy_2014_conditioning_study3",
 4:"Expectancy (high-low pain cue) / placebo vs control (Study 3)",
 5:"Representation of aversive prediction errors in the human periaqueductal gray",
 6:"Nat Neurosci 2014;17(11):1607-12. doi:10.1038/nn.3832",
 7:"Conditioning / Placebo",
 8:"Cue-based expectancy + placebo-analgesic cream (Study 3, pain-learning task)",
 9:"Heat (thermal)",10:"Left volar forearm",11:"Tonic (~11 s; 7.5 s plateau)",
 12:"50",13:"Healthy adults",14:"Expectancy / placebo modulation",
 15:"Placebo cream (placebo & control runs)",16:"Yes",
 17:"ie2 = Roy 2014 Nat Neurosci Study 3 (n=50), CU Boulder. Placebo-cream + 2-cue (46-47 / 47-48 C) expectancy design; fixed temps. Relabelled from prior blank ie / Jepma_2018_study1. Verify which contrast maps to the NPS/SIIPS score.",
 18:"Thermal",19:"Fixed (46/47/48 C)",20:"54% (27/50)"}
for c,v in vals.items(): ws.cell(r,c,v)

# ---- 1b. MPA2 heat & pressure (Ceko 2026) ----
ceko_cite="Ceko M, Magal N, Matthewson G, Bo K, Dehghani A, Spiegel D, Kober H, Wager TD. (unpublished, 2026). GitHub: canlab/2025_Ceko_MPA2_Nonreact"
r=idx["MPA2_heat_regulation"]
for c,v in {5:"Mindful non-reactivity alters the neural construction of negative affect",6:ceko_cite,
 7:"Mindfulness",8:"Mindful non-reactive acceptance (NR) vs control (CON) - heat pain",
 9:"Heat (thermal)",10:"Thenar eminence (hand); Medoc ATS Pathway",11:"Tonic (verify)",
 12:"55",13:"Healthy adults",14:"Non-reactive acceptance modulation",15:"N.A.",16:"Yes",
 17:"MPA2 = Ceko 2026 (unpublished). NR vs CON during heat pain; part of multimodal (pressure/heat/auditory/visual) study, 4 fixed intensity levels.",
 18:"Thermal",19:"Fixed (4 levels)",20:"44% (24/55)"}.items(): ws.cell(r,c,v)
r=idx["MPA2_pressure_regulation"]
for c,v in {5:"Mindful non-reactivity alters the neural construction of negative affect",6:ceko_cite,
 7:"Mindfulness",8:"Mindful non-reactive acceptance (NR) vs control (CON) - pressure pain",
 9:"Mechanical (pressure)",10:"Left thumbnail (pneumatic, 4-7 kg/cm2)",11:"Tonic (verify)",
 12:"55",13:"Healthy adults",14:"Non-reactive acceptance modulation",15:"N.A.",16:"Yes",
 17:"MPA2 = Ceko 2026 (unpublished). NR vs CON during pressure pain; same session/sample as MPA2 heat row.",
 18:"Non-thermal",19:"Fixed (4 levels)",20:"44% (24/55)"}.items(): ws.cell(r,c,v)

# ---- 3. add 'Females (n)' column (col 21) ----
col=21
c=ws.cell(1,col,"Females (n)"); c.fill=HFILL; c.font=HFONT
c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BD
ws.column_dimensions[get_column_letter(col)].width=12
FEM={"Atlas_2010_expectancy":9,"Atlas_2012_remifentanil":11,"Brascher_2016_control":10,
 "Becker_2017_reward":12,"Bingel_2011_remifentanil_posExpectancy":7,"Woo_2015_selfregulation":22,
 "Habermann_2025_control":"(verify)","Habermann_2025_predictability":"(verify)",
 "Roy_2014_conditioning_study3":27,"Woo_2017_expectancy":16,"Woo_2017_perceivedControl":16,
 "Jepma_2018_expectancy_study2":"~17","Koban_2019_socialExpectancy":20,"Kober_2019_mindfulAcceptance":5,
 "LopezSola_2018_prosocialMeaning":29,"LopezSola_2019_handholding":30,"MPA2_heat_regulation":24,
 "MPA2_pressure_regulation":24,"BotvinikNezer_2024_placebo":232,"Roy_2009_emotionModulation":7,
 "Steininger_2025_natureExposure":24,"Kim_2024_cueIntegration":24,"Tinnermann_2022_remifentanil":0,
 "Riegner_2023_mindfulness":20,"Bingel_2006_placebo":4,"Bingel_2011_placebo":7,"Choi_2011_placebo":0,
 "Eippert_2009_placebo":0,"Ellingsen_2013_placebo":10,"Elsenbruch_2012_placebo":18,"Freeman_2015_placebo":12,
 "Geuter_2013_placebo":"(verify)","Kong_2006_placebo":7,"Kong_2009_placebo":"~6","Lui_2010_placebo":18,
 "Schenk_2014_placebo":15,"Theysohn_2014_placebo":15,"Wager_2004_study1_placebo":"(verify)",
 "Wager_2004_study2_placebo":"(verify)","Wrobel_2014_placebo":23}
studies_order=[]
for r in range(2,ws.max_row+1):
    lbl=ws.cell(r,3).value
    if not lbl: continue
    studies_order.append(lbl)
    cc=ws.cell(r,col,FEM.get(lbl,"(verify)")); cc.font=CFONT; cc.border=BD; cc.alignment=Alignment(vertical="top")
ws.auto_filter.ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

# ---- 2. rebuild Scanning to match Studies order 1:1 ----
sc=wb["Scanning parameters"]
harvest={}
for r in range(2,sc.max_row+1):
    lbl=sc.cell(r,1).value
    if lbl: harvest[lbl]=[sc.cell(r,cc).value for cc in range(2,15)]
harvest["Roy_2014_conditioning_study3"]=["CU Boulder (CINC)","Siemens Trio","3 T","Echo-planar (EPI)",
 "1300","25","(n/r)","3.4 x 3.4 x 3.0 mm","26","220 mm","(per main exp.; verify)","Verified",
 "Roy 2014 Nat Neurosci Study 3 (n=50). CINC Boulder Trio. Preproc per main experiment (main = 1.5T Columbia spiral); resampled 2 mm."]
mpa=["CU Boulder (Intermountain Neuroimaging Consortium)","Siemens MAGNETOM Prisma","3 T",
 "Multiband EPI (MB=8)","460","27.2","44","2.7 x 2.7 x 2.7 mm","56 (interleaved asc.)","220 mm; 64x64","6 mm","Verified",
 "Ceko 2026 (MPA2, unpublished). Same Boulder Prisma protocol as Botvinik-Nezer 2024 / Kim 2024. First 15 vols discarded; resample 2 mm; SPM8; 6 runs x 7.17 min per scan (CON, NR)."]
harvest["MPA2_heat_regulation"]=list(mpa)
harvest["MPA2_pressure_regulation"]=list(mpa)
alias={"Habermann_2025_control":"Habermann_2025_control / _predictability",
 "Habermann_2025_predictability":"Habermann_2025_control / _predictability",
 "Woo_2017_expectancy":"Woo_2017_expectancy / _perceivedControl",
 "Woo_2017_perceivedControl":"Woo_2017_expectancy / _perceivedControl",
 "Jepma_2018_expectancy_study2":"Jepma_2018_expectancy"}
def params_for(lbl):
    if lbl in harvest: return harvest[lbl]
    if lbl in alias and alias[lbl] in harvest: return list(harvest[alias[lbl]])
    return ["(verify)"]*11+["NOT RETRIEVED","(row added to match Studies; fill params)"]

for r in range(sc.max_row,1,-1): sc.delete_rows(r)
for i,lbl in enumerate(studies_order,2):
    row=[lbl]+params_for(lbl)
    for j,v in enumerate(row,1):
        cc=sc.cell(i,j,v); cc.font=CFONT; cc.border=BD; cc.alignment=Alignment(vertical="top",wrap_text=True)
    st=sc.cell(i,13).value
    sc.cell(i,13).fill = GREEN if st=="Verified" else (GREY if str(st).startswith("N/A") else ORANGE)
sc.auto_filter.ref=f"A1:{get_column_letter(sc.max_column)}{len(studies_order)+1}"
sc.freeze_panes="B2"

wb.save(FN)
print("Saved. Studies rows:",len(studies_order),"| Scanning data rows:",sc.max_row-1)
wb2=openpyxl.load_workbook(FN, read_only=True)
sl=[r[2] for r in wb2["Studies"].iter_rows(min_row=2,values_only=True) if r[2]]
cl=[r[0] for r in wb2["Scanning parameters"].iter_rows(min_row=2,values_only=True) if r[0]]
mism=[(a,b) for a,b in zip(sl,cl) if a!=b]
print("1:1 label match:", "YES" if sl==cl else "NO "+str(mism[:5]))
print("counts: studies",len(sl),"scanning",len(cl))
