# -*- coding: utf-8 -*-
import io, sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN="Study_summary_completed.xlsx"

# New Label -> "% female (n F / N)"  string
SEX={
"Atlas_2010_expectancy":"50% (9/18)","Atlas_2012_remifentanil":"52% (11/21)",
"Becker_2017_reward":"50% (12/24)","Bingel_2011_remifentanil_posExpectancy":"32% (7/22)",
"Woo_2015_selfregulation":"67% (22/33)","Habermann_2025_control":"(verify)",
"Habermann_2025_predictability":"(verify)","Woo_2017_expectancy":"55% (16/29)",
"Woo_2017_perceivedControl":"55% (16/29)","Jepma_2018_expectancy_study2":"50% (Study 2; ~17/34)",
"Koban_2019_socialExpectancy":"56% (20/36)","Kober_2019_mindfulAcceptance":"29% (5/17)",
"LopezSola_2018_prosocialMeaning":"100% (29/29)","LopezSola_2019_handholding":"100% (30/30)",
"BotvinikNezer_2024_placebo":"59% (232/395)","Roy_2009_emotionModulation":"54% (7/13)",
"Steininger_2025_natureExposure":"49% (24/49)","Kim_2024_cueIntegration":"50% (24/48)",
"Tinnermann_2022_remifentanil":"0% (all male)","Riegner_2023_mindfulness":"50% (20/40)",
"Bingel_2006_placebo":"21% (4/19)","Bingel_2011_placebo":"32% (7/22)",
"Choi_2011_placebo":"0% (all male)","Eippert_2009_placebo":"0% (all male)",
"Ellingsen_2013_placebo":"36% (10/28)","Elsenbruch_2012_placebo":"50% (18/36)",
"Freeman_2015_placebo":"50% (12/24)","Geuter_2013_placebo":"(verify)",
"Kong_2006_placebo":"44% (7/16 completed)","Kong_2009_placebo":"50% (~6/12, balanced)",
"Lui_2010_placebo":"58% (18/31)","Schenk_2014_placebo":"47% (15/32)",
"Theysohn_2014_placebo":"50% (15/30)","Wager_2004_study1_placebo":"(verify)",
"Wager_2004_study2_placebo":"(verify)","Wrobel_2014_placebo":"61% (23/38)",
"MPA2_heat_regulation":"(verify)","MPA2_pressure_regulation":"(verify)",
"Jepma_2018_expectancy_study1":None,   # intentionally blank
}

HFILL=PatternFill("solid",fgColor="1F4E78"); HFONT=Font(name="Arial",bold=True,color="FFFFFF",size=10)
CFONT=Font(name="Arial",size=10); THIN=Side(style="thin",color="D9D9D9"); BD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

wb=openpyxl.load_workbook(FN); ws=wb["Studies"]
col=ws.max_column+1
c=ws.cell(1,col,"% female (n F / N)")
c.fill=HFILL; c.font=HFONT; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BD
ws.column_dimensions[get_column_letter(col)].width=20

miss=[]
for r in range(2,ws.max_row+1):
    lbl=ws.cell(r,3).value
    val=SEX.get(lbl,"(verify)")
    if lbl not in SEX: miss.append(lbl)
    cc=ws.cell(r,col,val); cc.font=CFONT; cc.border=BD; cc.alignment=Alignment(vertical="top",wrap_text=True)
ws.auto_filter.ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
wb.save(FN)
print("Added '% female' col at",col)
if miss: print("labels not in map:",miss)

wb2=openpyxl.load_workbook(FN, read_only=True); s=wb2["Studies"]
print()
for row in s.iter_rows(min_row=2, values_only=True):
    if row[2]: print(f"  {str(row[2])[:34]:35} {row[-1]}")
