# -*- coding: utf-8 -*-
import io, sys
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN = "Study_summary_completed.xlsx"

# New Label -> {col: value}.  col 6 = Citation/DOI, col 18 = Notes (append-style)
FIXES = {
"Elsenbruch_2012_placebo": {
  6: "Kotsis V, Benson S, Bingel U, Forsting M, Schedlowski M, Gizewski ER, Elsenbruch S. Neurogastroenterol Motil 2012;24(10):935-e463. doi:10.1111/j.1365-2982.2012.01968.x",
  18: "CITATION CORRECTED: journal was wrong (was NeuroImage); actually Neurogastroenterol Motil. First author Kotsis V (Elsenbruch senior). Visceral (rectal distension); expectancy; saline."},
"Theysohn_2014_placebo": {
  6: "Theysohn N, Schmid J, Icenhour A, Mewes C, Forsting M, Gizewski ER, Schedlowski M, Elsenbruch S. Neurogastroenterol Motil 2014;26(12):1745-53. doi:10.1111/nmo.12454",
  18: "CITATION COMPLETED from PDF (DOI 10.1111/nmo.12454). Visceral placebo; sex-differences study (Essen)."},
"Choi_2011_placebo": {
  6: "Choi JC, Yi DJ, Han BS, Lee PH, Kim JH, Kim BH. NeuroReport 2011;22(9):419-23. doi:10.1097/WNR.0b013e32834601c9",
  18: "CITATION COMPLETED from PDF (DOI 10.1097/WNR.0b013e32834601c9). Correct paper now in folder."},
}

wb = openpyxl.load_workbook(FN)
ws = wb["Studies"]
labels = {ws.cell(r,3).value: r for r in range(2, ws.max_row+1)}
for label, edits in FIXES.items():
    r = labels.get(label)
    if not r:
        print("  !! not found:", label); continue
    for col, val in edits.items():
        ws.cell(r, col, val)
    print("  fixed:", label, "->", ws.cell(r,6).value[:70])
wb.save(FN)
print("Saved", FN)
