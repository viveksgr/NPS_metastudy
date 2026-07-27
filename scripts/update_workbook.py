# -*- coding: utf-8 -*-
# update_workbook.py
# Fills the "Scanning parameters" and "Studies" sheets of
# Study_summary_completed.xlsx from parameters extracted from the PDFs.
import io, sys
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN = "Study_summary_completed.xlsx"

GREEN  = PatternFill("solid", fgColor="E2EFDA")  # Verified
ORANGE = PatternFill("solid", fgColor="FCE4D6")  # partial / not retrieved
GREY   = PatternFill("solid", fgColor="D9D9D9")  # N/A (no fMRI)

# scan params keyed by "New Label" (column A of Scanning parameters sheet)
# order: Site, Scanner, Field, Sequence, TR, TE, Flip, Voxel, Slices, FOV, Smoothing, Status, Note
SCAN = {
"Atlas_2012_remifentanil": ["Columbia (Wager)","GE Signa Twin Speed Excite HD","1.5 T",
  "Echo-planar (EPI)","2000","34","(n/r)","3.5 x 3.5 x 4.0 mm","28","224 mm; 64x64","8 mm","Verified",
  "J Neurosci 2012. Verified from full text. Behavioural n=19 (imaging subsets n=14-21)."],

"Brascher_2016_control": ["McGill / MNI (McConnell BIC)","Siemens Trio (32-ch)","3 T",
  "Gradient-echo EPI (whole brain)","2620","30","90","3.5 x 3.5 x 3.5 mm","44 (interleaved)","224 x 224 mm; 64x64","5 mm","Verified",
  "J Neurosci 2016. 441 volumes. T1 MPRAGE 1 mm iso. Final N=23."],

"Bingel_2011_remifentanil_posExpectancy": ["Hamburg / Oxford (Bingel/Tracey)","Varian / Siemens 3-T system","3 T",
  "T2* EPI (brain + brainstem)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","Verified (partial)",
  "Sci Transl Med 2011. 4-channel head coil. TR/TE/voxel etc. in Supplementary Methods (not in main PDF)."],

"Jepma_2018_expectancy": ["Boulder (Wager), CINC","Siemens Trio","3 T",
  "Echo-planar (EPI), SENSE x2","1300","25","(n/r)","3.4 x 3.4 x 3.0 mm","26","220 mm","6 mm","Verified",
  "Nat Hum Behav 2018. Study 2 fMRI dataset (n=34). Parallel imaging SENSE factor 2."],

"LopezSola_2018_prosocialMeaning": ["Boulder (Wager)","Siemens TrioTim","3 T",
  "T2* EPI (GRAPPA)","1300","25","50","3.4 mm isomorphic","26 (interleaved)","220 mm","8 mm","Verified",
  "Psychosom Med 2018. N=29 women. T1 MPRAGE 1 mm iso. Resampled to 2 mm."],

"LopezSola_2019_handholding": ["Boulder (Wager)","Siemens TrioTim","3 T",
  "T2* EPI (GRAPPA)","1300","25","50","3.4 mm isomorphic","26 (interleaved)","220 mm","8 mm","Verified",
  "Pain 2019. N=30 women. Same Boulder protocol as Lopez-Sola 2018."],

"Roy_2009_emotionModulation": ["Montreal (Rainville), CRIUGM","Siemens Trio (CP head coil)","3 T",
  "T2* gradient EPI (BOLD)","3000 (500 ms intervol. delay)","30","90","3.44 x 3.44 x 5 mm","41","(n/r); 64x64 matrix","(n/r)","Verified",
  "PNAS 2009. Combined brain + spinal cord; brain protocol listed. T1 1x1x1.1 mm."],

"Riegner_2023_mindfulness": ["UCSD (Zeidan)","Siemens Magnetom Skyra (32-ch)","3 T",
  "BOLD EPI","2000","25","75","4.0 x 4.0 x 4.0 mm","35 (no gap)","(n/r)","5 mm","Verified",
  "Pain 2023. N=40 (20/group). T1: 1 mm iso, 192 slices, flip 9, GRAPPA 2."],

"Bingel_2006_placebo": ["Hamburg (Buchel)","Siemens Vision","1.5 T",
  "Gradient-echo EPI (T2*)","2600","40","90","~3.3 x 3.3 mm in-plane; 3 mm slice + 1 mm gap","32 (axial)","210 x 210 mm; 64x64","8 mm","Verified",
  "Pain 2006. Laser pain. T1 3D FLASH 1 mm. In-plane ~3.3 mm (FOV210/64)."],

"Bingel_2011_placebo": ["Hamburg / Oxford","Varian / Siemens 3-T system","3 T",
  "T2* EPI (brain + brainstem)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","(in suppl.)","Verified (partial)",
  "Sci Transl Med 2011 (same study as Bingel_2011 remifentanil row). 4-channel coil. Params in Suppl Methods."],

# Choi: downloaded PDF is the WRONG paper - leave NOT RETRIEVED, flag it
"Choi_2011_placebo": ["Korea (Choi)","","","","","","","","","","","NOT RETRIEVED",
  "WRONG PDF in folder: it is Necka et al. 2025 (J Neurosci, NIH dual 3T Skyra, multi-echo), NOT Choi 2011. Original Choi 2011 (NeuroReport, electrical, saline) still not located - verify citation."],

"Ellingsen_2013_placebo": ["Oslo / Gothenburg","Philips Achieva (whole-body)","3 T",
  "EPI (details in SI)","(in SI)","(in SI)","(in SI)","(in SI)","(in SI)","(in SI)","(in SI)","Verified (partial)",
  "PNAS 2013. Scanner+field from main text; TR/TE/voxel in SI Materials & Methods (not in downloaded main PDF)."],

"Elsenbruch_2012_placebo": ["Essen (Elsenbruch)","Siemens Sonata","1.5 T",
  "Echo-planar (BOLD)","3100","50","90","~3.75 x 3.75 mm in-plane; 3 mm slice + 0.3 mm gap","34 (transversal)","240 mm; 64 matrix","9 mm","Verified",
  "NeuroImage 2012. Visceral (rectal distension). T1 3D FLASH 1.5 mm."],

"Freeman_2015_placebo": ["MGH (Kong/Gollub)","Siemens TIM Trio (3-axis gradient head coil)","3 T",
  "Echo-planar (EPI)","2000","40","90","3.13 x 3.13 x 4 mm (4 mm + 1 mm skip)","30 (axial)","(n/r); in-plane 3.13 mm","8 mm","Verified",
  "NeuroImage 2015. Same MGH protocol as Kong 2006/2009."],

"Geuter_2013_placebo": ["Hamburg (Buchel)","Siemens Trio (32-ch)","3 T",
  "T2* EPI","2580","26","80","2 x 2 x 2 mm (1 mm gap)","42 (transversal)","220 x 220 mm","6 mm","Verified",
  "NeuroImage 2013. Slices tilted ~30 deg (OFC SNR)."],

"Kong_2009_placebo": ["MGH (Kong/Gollub)","Siemens Allegra (head-only) -> TIM Trio (mid-study upgrade)","3 T",
  "Echo-planar (EPI)","2000","40","90","3.13 x 3.13 x 4 mm (4 mm + 1 mm skip)","30 (axial)","(n/r); in-plane 3.13 mm","8 mm","Verified",
  "NeuroImage 2009. Sham acupuncture. Scanner upgraded Allegra->TIM Trio mid-study; params held constant."],

"Lui_2010_placebo": ["Modena (Porro)","Philips Intera","3 T",
  "Gradient-echo EPI","3000","35","(n/r)","1.9 x 1.9 x 3.5 mm (0.5 mm gap)","30 (axial)","240 mm; matrix 80x80 (recon 128x128)","4 x 4 x 8 mm","Verified",
  "Pain 2010. Laser pain, sham TENS. T1 1x1x1 mm, 170 sagittal slices."],

"Schenk_2014_placebo": ["Hamburg (Buchel)","Siemens Magnetom TIM Trio (32-ch)","3 T",
  "Gradient EPI (GRAPPA PAT 2)","2580","26","(n/r)","2 x 2 x 2 mm (1 mm gap)","42 (transversal)","(n/r)","6 mm","Verified",
  "Pain 2014. Heat+capsaicin. Same Hamburg protocol as Geuter 2013. T1 MPRAGE 1 mm."],

"Theysohn_2014_placebo": ["Essen (Elsenbruch/Theysohn)","Siemens Magnetom Skyra (16-ch)","3 T",
  "T2* gradient EPI","2400","26","90","~2.55 mm in-plane; 3 mm slice","42","240 mm; matrix 94","8 mm","Verified",
  "Visceral (rectal distension). T1 MPRAGE TR1900/TE2.13/flip9/0.9 mm/192 slices."],

"Wager_2004_study1_placebo": ["Michigan (Wager)","(n/r in main text)","(n/r)",
  "Echo-planar (EPI)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","NOT RETRIEVED (partial)",
  "Science 2004 Study 1 (electrical shock). Study 1 used EPI; acquisition detail only in Supporting Online Material."],

"Wager_2004_study2_placebo": ["Princeton (Wager)","(n/r in main text)","(n/r)",
  "Spiral (gradient-echo)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","(in SOM)","NOT RETRIEVED (partial)",
  "Science 2004 Study 2 (heat). Study 2 used spiral imaging (more OFC dropout). Detail in Supporting Online Material."],

"Wrobel_2014_placebo": ["Hamburg (Bingel)","Siemens Magnetom Trio (12-ch)","3 T",
  "T2* gradient-echo EPI (GRAPPA PAT 2)","2600","25","90","2.0 x 2.0 mm in-plane; 1 mm gap","44 (transversal, descending)","208 x 208 mm; 104x104","8 mm","Verified",
  "Cortex 2014. Slice tilt 39 deg. T1 MPRAGE 1 mm."],

# Jepma 2015 is behavioural/SCR - no fMRI
"Jepma_2015_conceptualConditioning": ["Boulder (Wager)","N/A - no fMRI","N/A",
  "N/A (behavioural / skin-conductance study)","-","-","-","-","-","-","-","N/A (no fMRI)",
  "Psychol Sci 2015. Behavioural + SCR study, NO fMRI. 26 participants (30 tested). Should be excluded from scanning-parameter analysis."],
}

# Studies-sheet edits keyed by "New Label" (column C): {col_index: value}
#   col 6 = Citation/DOI, 13 = N (analyzed), 18 = Notes
STUDIES_EDITS = {
"Kim_2024_cueIntegration": {6: "Sci Adv 2024;10. doi:10.1126/sciadv.ado8230",
    18: "Internal name 'suhwan' = S. Gim (co-author); first author Kim J. DOI corrected adk7421 -> ado8230."},
"Steininger_2025_natureExposure": {6: "Nat Commun 2025;16. doi:10.1038/s41467-025-56870-2",
    18: "DOI corrected -56870-x -> -56870-2."},
"Atlas_2012_remifentanil": {13: "21 (imaging); 19 (behavioural)",
    18: "Year corrected 2013->2012. N from full text: 19 behavioural, imaging subsets n=14-21."},
"Brascher_2016_control": {13: "23",
    18: "Final sample N=23 (13 male). Original note: no subjective ratings for contrast."},
"LopezSola_2018_prosocialMeaning": {13: "29 (women)",
    18: "Overlaps social & cognitive-reappraisal. N=29 healthy women (verified from text)."},
"Riegner_2023_mindfulness": {13: "40 (20/group)",
    18: "First author Riegner (Zeidan lab). N=40, 20/group, 20M/20F, mean age 30 (verified)."},
"Jepma_2015_conceptualConditioning": {13: "26 (behavioural; NO fMRI)",
    18: "BEHAVIOURAL / skin-conductance study - NOT fMRI. 26 final (30 tested). No scanning parameters; consider excluding from imaging analyses."},
}

wb = openpyxl.load_workbook(FN)

# ---- Scanning parameters sheet ----
ws = wb["Scanning parameters"]
labels = {ws.cell(r,1).value: r for r in range(2, ws.max_row+1)}
n_upd = 0
for label, vals in SCAN.items():
    r = labels.get(label)
    if not r:
        print("  !! label not found in Scanning sheet:", label)
        continue
    for j, v in enumerate(vals, start=3):   # cols 3..15
        ws.cell(r, j, v)
    status = vals[-2]
    fill = GREEN if status == "Verified" else (GREY if status.startswith("N/A") else ORANGE)
    ws.cell(r, 14).fill = fill
    n_upd += 1
print(f"Scanning parameters: updated {n_upd} rows")

# ---- Studies sheet ----
ws2 = wb["Studies"]
labels2 = {ws2.cell(r,3).value: r for r in range(2, ws2.max_row+1)}
n2 = 0
for label, edits in STUDIES_EDITS.items():
    r = labels2.get(label)
    if not r:
        print("  !! label not found in Studies sheet:", label)
        continue
    for col, val in edits.items():
        ws2.cell(r, col, val)
    n2 += 1
print(f"Studies: updated {n2} rows")

# ---- refresh README counts if present ----
if "README" in wb.sheetnames:
    rm = wb["README"]
    ver = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,14).value == "Verified")
    part = sum(1 for r in range(2, ws.max_row+1) if str(ws.cell(r,14).value).startswith(("Verified (partial)","NOT RETRIEVED")))
    for r in range(1, rm.max_row+1):
        a = rm.cell(r,1).value
        if a == "Verified papers":
            rm.cell(r,2, str(ver))
        elif a and a.startswith("Not retrieved"):
            rm.cell(r,2, str(part))

wb.save(FN)
print("Saved", FN)

# report final status counts
vcount = {}
for r in range(2, ws.max_row+1):
    s = ws.cell(r,14).value
    vcount[s] = vcount.get(s,0)+1
print("\nFinal Scanning-parameter status counts:")
for k,v in sorted(vcount.items(), key=lambda x:-x[1]):
    print(f"  {v:3}  {k}")
