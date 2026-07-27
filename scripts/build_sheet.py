# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD = ["Published","Original Label","New Label (Lastname_year_contrast)","Contrast",
        "Study Title","Citation / DOI","Type of Intervention (original)",
        "Type of Intervention (validated, broad)","Mechanism / subtype","Pain Modality",
        "Body Site","Pain Temporality","N (analyzed)","Population","Modulation Direction",
        "Placebo Conditioning","Analgesic Effect (-pain rating)","Notes"]

R = []
def add(*vals): R.append(list(vals))

add("Yes","atlas_2010_Hi_v_Low","Atlas_2010_expectancy","Low-expectancy cue - high-expectancy cue (matched heat)",
"Brain mediators of predictive cue effects on perceived pain","J Neurosci 2010;30(39):12964-77. doi:10.1523/JNEUROSCI.0057-10.2010",
"Conditioning","Expectancy / Conditioning","Predictive cue-based expectancy","Heat (thermal)","Left volar forearm","Tonic (~10 s)",
"19 (18 analyzed)","Healthy adults","Expectancy analgesia (low cue)","N.A.","Yes","")

add("Yes","atlas_2013_nodrug_v_drug","Atlas_2012_remifentanil","Remifentanil - no remifentanil (drug analgesia)",
"Dissociable influences of opiates and expectations on pain","J Neurosci 2012;32(23):8053-64. doi:10.1523/JNEUROSCI.0383-12.2012",
"Remifentanil","Opioid (pharmacological)","mu-opioid agonist (remifentanil) x open/hidden expectancy","Heat (thermal)","Forearm","Tonic",
"~19 (verify)","Healthy adults","Opioid analgesia","N.A.","Yes","Year corrected 2013->2012. N paywalled - verify.")

add("Yes","becker_control_2016","Brascher_2016_control","Controllable - uncontrollable pain (control-induced analgesia)",
"Different brain circuitries mediating controllable and uncontrollable pain","J Neurosci 2016;36(18):5013-25. doi:10.1523/JNEUROSCI.1954-15.2016",
"Control","Perceived control","Self-adjusted (controllable) vs yoked (uncontrollable) thermal","Heat (thermal)","Hand","Tonic (continuous self-adjust)",
"(verify)","Healthy adults","Control-induced analgesia","N.A.","No","Original note: no subjective ratings for contrast. N paywalled - verify.")

add("Yes","becker_2017_loss_v_win","Becker_2017_reward","Monetary win - loss (reward-induced analgesia)",
"Orbitofrontal cortex mediates pain inhibition by monetary reward","SCAN 2017;12(4):651-61. doi:10.1093/scan/nsw173",
"Cognitive","Cognitive-affective (reward)","Monetary reward (wheel-of-fortune) during pain","Heat (thermal)","Hand/arm","Phasic",
"24","Healthy adults","Reward-induced analgesia","N.A.","Yes","")

add("Yes","bingel_2011_remi","Bingel_2011_remifentanil_posExpectancy","Positive expectancy + remifentanil vs hidden/negative expectancy",
"The effect of treatment expectation on drug efficacy: imaging the analgesic benefit of the opioid remifentanil","Sci Transl Med 2011;3(70):70ra14. doi:10.1126/scitranslmed.3001244",
"Remifentanil","Opioid (pharmacological) x expectancy","Remifentanil under positive/negative treatment expectancy","Heat (thermal)","Left hand","Tonic (~20 s)",
"22","Healthy adults","Opioid x expectancy analgesia","N.A.","Yes","")

add("Yes","Woo_2015_bmrk3_up_v_down","Woo_2015_selfregulation","Regulate-down - regulate-up (cognitive self-regulation)",
"Distinct brain systems mediate the effects of nociceptive input and self-regulation on pain","PLoS Biol 2015;13(1):e1002036. doi:10.1371/journal.pbio.1002036",
"Cognitive","Cognitive self-regulation","Imagery-based up/down self-regulation","Heat (thermal)","Left forearm","Tonic (~11 s)",
"33","Healthy adults","Self-regulation analgesia (down)","N.A.","Yes","")

add("Yes","habermann_control","Habermann_2025_control","Controllable - predictable pain",
"Controllability changes pain perception by increasing the precision of expectations","Nat Commun 2025;16. doi:10.1038/s41467-025-66038-7",
"Control","Perceived control","Control vs (matched) predict condition","Heat (thermal)","Arm/hand","Phasic/Tonic",
"59 (55 analyzed)","Healthy adults","Control-induced analgesia","N.A.","Yes","Control & predict are two contrasts of the same study.")

add("Yes","habermann_predict","Habermann_2025_predictability","Predictable - unpredictable pain",
"Controllability changes pain perception by increasing the precision of expectations","Nat Commun 2025;16. doi:10.1038/s41467-025-66038-7",
"Cognitive","Expectancy / Conditioning","Predictability of stimulus (precision of expectation)","Heat (thermal)","Arm/hand","Phasic/Tonic",
"59 (55 analyzed)","Healthy adults","Predictability modulation","N.A.","Yes","Reclassified from 'Cognitive' -> predictability/expectancy.")

add("Yes","ie","Jepma_2015_conceptualConditioning","High-cue - low-cue (conceptual/symbolic conditioning)",
"Conceptual conditioning: mechanisms mediating conditioning effects on pain","Psychol Sci 2015;26(11):1728-39. doi:10.1177/0956797615597658",
"Conditioning","Expectancy / Conditioning","Conceptual (symbolic) conditioning of cues","Heat (thermal)","Forearm (multiple skin sites)","Phasic (~1 s peak)",
"(verify)","Healthy adults","Cue-based modulation (low cue analgesia)","N.A.","Yes","N paywalled - verify.")

add("Yes","ilcp_exp","Woo_2017_expectancy","High - low expectancy cue",
"Quantifying cerebral contributions to pain beyond nociception","Nat Commun 2017;8:14211. doi:10.1038/ncomms14211",
"Conditioning","Expectancy / Conditioning","Cue-based expectancy (SIIPS-1 dataset)","Heat (thermal)","Forearm/leg","Tonic",
"multi-study (137 train / 46 test)","Healthy adults","Expectancy modulation","N.A.","Yes","SIIPS-1 paper pools several datasets; ilcp = one expectancy dataset. Verify subset N.")

add("Yes","ilcp_ctrl","Woo_2017_perceivedControl","Controllable - uncontrollable",
"Quantifying cerebral contributions to pain beyond nociception","Nat Commun 2017;8:14211. doi:10.1038/ncomms14211",
"Control","Perceived control","Perceived control (SIIPS-1 dataset)","Heat (thermal)","Forearm/leg","Tonic",
"multi-study (137 train / 46 test)","Healthy adults","Control-induced analgesia","N.A.","Yes","SIIPS-1 paper; ilcp = one control dataset. Verify subset N.")

add("Yes","jepma_2018_high_v_low","Jepma_2018_expectancy","High - low expectancy cue (self-reinforcing expectancy)",
"Behavioural and neural evidence for self-reinforcing expectancy effects on pain","Nat Hum Behav 2018;2:838-855. doi:10.1038/s41562-018-0455-8",
"Conditioning","Expectancy / Conditioning","Self-reinforcing cue-based expectancy","Heat (thermal)","Lower leg (fMRI)","Phasic (~1 s peak)",
"34 (fMRI); 28 behavioral","Healthy adults","Cue-based modulation (low cue analgesia)","N.A.","Yes","")

add("Yes","koban_2019_social_pain","Koban_2019_socialExpectancy","Social expectation effect (high - low social cue) on pain",
"Different brain networks mediate the effects of social and conditioned expectations on pain","Nat Commun 2019;10:4096. doi:10.1038/s41467-019-11934-y",
"Social","Social","Social (observed ratings) vs conditioned cues","Heat (thermal)","Right calf/leg","Phasic",
"38 (36 analyzed)","Healthy adults","Social-expectation modulation","N.A.","Yes","")

add("Yes","kober_2019_mindful_acceptance","Kober_2019_mindfulAcceptance","Mindful-acceptance - react naturally (pain)",
"Let it be: mindful acceptance down-regulates pain and negative emotion","SCAN 2019;14(11):1147-1158. doi:10.1093/scan/nsz104",
"Mindfulness","Mindfulness","Mindful-acceptance regulation","Heat (thermal)","Forearm","Tonic",
"16","Healthy adults (meditation-naive)","Mindfulness analgesia","N.A.","Yes","")

add("Yes","lopezsola_2018_pain_meaning","LopezSola_2018_prosocialMeaning","Prosocial meaning - control (pain reappraisal)",
"Transforming pain with prosocial meaning: a functional magnetic resonance imaging study","Psychosom Med 2018;80(9):814-825. doi:10.1097/PSY.0000000000000609",
"Social","Cognitive-affective (meaning) / Social","Reframing pain as prosocial/meaningful","Heat (thermal)","Forearm/hand","Tonic",
"~28 (verify)","Healthy adults","Meaning-based analgesia","N.A.","Yes","Overlaps social & cognitive-reappraisal. N paywalled - verify.")

add("Yes","lopezsola_2019_handholding","LopezSola_2019_handholding","Partner hand-holding - inert device (social touch)",
"Brain mechanisms of social touch-induced analgesia in females","Pain 2019;160(9):2072-2085. doi:10.1097/j.pain.0000000000001599",
"Social","Social (social touch)","Romantic-partner hand-holding (social support)","Heat (thermal)","Forearm","Tonic",
"30","Healthy females","Social-touch analgesia","N.A.","Yes","")

add("No","MPA2_heat_reg-exp","MPA2_heat_regulation","Regulate - experience (heat)",
"Unpublished CANlab dataset (MPA2)","(unpublished)",
"Mindfulness","Cognitive self-regulation (uncertain)","Self-regulation / mindful regulation of heat pain","Heat (thermal)","(not reported)","(not reported)",
"(unpublished - verify)","Healthy adults","Self-regulation analgesia","N.A.","Yes","Unpublished. Broad group uncertain (regulation vs mindfulness).")

add("No","MPA2_pressure_reg-exp","MPA2_pressure_regulation","Regulate - experience (pressure)",
"Unpublished CANlab dataset (MPA2)","(unpublished)",
"Mindfulness","Cognitive self-regulation (uncertain)","Self-regulation / mindful regulation of pressure pain","Mechanical (pressure)","(not reported)","(not reported)",
"(unpublished - verify)","Healthy adults","Self-regulation analgesia","N.A.","Yes","Unpublished. Broad group uncertain.")

add("Yes","paingen_placebo","BotvinikNezer_2024_placebo","Placebo - control",
"Placebo treatment affects brain systems related to affective and cognitive processes, but not nociceptive pain","Nat Commun 2024;15:6017. doi:10.1038/s41467-024-50103-8",
"Placebo","Placebo (+ conditioning)","Placebo cream with conditioning/expectancy","Heat (thermal)","Forearm/leg","Tonic",
"392","Healthy adults","Placebo analgesia","Conditioned + expectancy","Yes","Large pre-registered sample.")

add("Yes","roy_emomod_2009","Roy_2009_emotionModulation","Negative - positive/neutral emotional context on pain",
"Cerebral and spinal modulation of pain by emotions","PNAS 2009;106(49):20900-5. doi:10.1073/pnas.0904706106",
"Cognitive","Cognitive-affective (emotion)","Emotional picture context (IAPS) modulation","Electrical","Right ankle (sural nerve)","Phasic (30 ms train)",
"12 (fMRI)","Healthy adults","Emotion-induced analgesia (positive context)","N.A.","Yes","")

add("Yes","nature_exposure","Steininger_2025_natureExposure","Nature exposure - control on pain",
"Nature exposure induces analgesic effects by acting on nociception-related neural processing","Nat Commun 2025;16. doi:10.1038/s41467-025-56870-2",
"Cognitive","Cognitive (environmental/attention)","Virtual nature exposure","Electrical","Hand","Phasic (1000 ms shock)",
"48 (preregistered; 49 reported)","Healthy adults","Nature-exposure analgesia","N.A.","Yes","")

add("Yes","suhwan_semi","Kim_2024_cueIntegration","High - low cue (cue-stimulus expectancy integration)",
"A computational mechanism of cue-stimulus integration for pain in the brain","Sci Adv 2024;10. doi:10.1126/sciadv.ado8230",
"Cognitive","Expectancy / Conditioning","Cue-stimulus (Bayesian) integration","Heat (thermal)","Left forearm","Tonic (~12.5 s)",
"84 recruited (final - verify)","Healthy adults","Cue-based modulation","N.A.","Yes","Internal name 'suhwan' = S. Gim (co-author); first author Kim J. Final analyzed N verify.")

add("Yes","tinnermann_remi","Tinnermann_2022_remifentanil","Remifentanil - saline (opioid analgesia)",
"Opioid analgesia alters corticospinal coupling along the descending pain system in healthy participants","eLife 2022;11:e74293. doi:10.7554/eLife.74293",
"Remifentanil","Opioid (pharmacological)","Remifentanil infusion","Heat (thermal)","Left volar forearm","Tonic (~15 s)",
"78","Healthy males","Opioid analgesia","N.A.","Yes","")

add("Yes","zeidan_mindfulness","Riegner_2023_mindfulness","Mindfulness meditation - control (meditation analgesia)",
"Disentangling self from pain: mindfulness meditation-induced pain relief is driven by thalamic-default mode network decoupling","Pain 2023;164(2):280-291. doi:10.1097/j.pain.0000000000002731",
"Mindfulness","Mindfulness","Mindfulness meditation","Heat (thermal)","Calf/leg","Tonic",
"~40 (verify)","Healthy adults","Meditation analgesia","N.A.","Yes","First author Riegner (Zeidan lab). N - verify.")

def placebo(orig,new,title,cite,grp,modality,site,temp,N,ptype,note=""):
    broad = grp.split("|")[1]
    contrast = "Placebo (conditioned) - control" if "conditioning" in broad.lower() else "Placebo (expectancy) - control"
    add("Yes",orig,new,contrast,title,cite,grp.split("|")[0],broad,"Placebo treatment ("+ptype+")",
        modality,site,temp,N,"Healthy adults","Placebo analgesia",ptype,"Yes",note)

placebo("bingel_2006_placebo","Bingel_2006_placebo",
"Mechanisms of placebo analgesia: rACC recruitment of a subcortical antinociceptive network","Pain 2006;120(1-2):8-15. doi:10.1016/j.pain.2005.08.027",
"Placebo+C|Placebo + conditioning","Laser","Dorsum of hand","Phasic (laser)","19","Conditioned; inert cream")

placebo("bingel_2011_placebo","Bingel_2011_placebo",
"The effect of treatment expectation on drug efficacy: imaging the analgesic benefit of the opioid remifentanil","Sci Transl Med 2011;3(70):70ra14. doi:10.1126/scitranslmed.3001244",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Left hand","Tonic","22","Conditioned; remifentanil infusion","Placebo/expectancy condition of the remifentanil study.")

placebo("choi_2011_placebo","Choi_2011_placebo",
"Placebo effects on analgesia related to testosterone and premotor activation","Choi JC et al. 2011 (NeuroReport)",
"Placebo+C|Placebo + conditioning","Electrical","Hand","Phasic","15","Conditioned; saline infusion","Exact citation/journal - verify.")

placebo("Eippert_2009_placebo","Eippert_2009_placebo",
"Activation of the opioidergic descending pain control system underlies placebo analgesia","Neuron 2009;63(4):533-43. doi:10.1016/j.neuron.2009.07.014",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Forearm","Tonic","40","Conditioned; inert cream")

placebo("Ellingsen_2013_placebo","Ellingsen_2013_placebo",
"Placebo improves pleasure and pain through opposite modulation of sensory processing","PNAS 2013;110(44):17993-8. doi:10.1073/pnas.1305050110",
"Placebo|Placebo (expectancy)","Heat (thermal)","Forearm","Tonic","28","Expectancy; inert nasal spray")

placebo("Elsenbruch_2012_placebo","Elsenbruch_2012_placebo",
"Perceived treatment group affects behavioral and neural responses to visceral pain in a deceptive placebo study","NeuroImage 2012;63(1):485-93. doi:10.1016/j.neuroimage.2012.07.002",
"Placebo|Placebo (expectancy)","Visceral (rectal distension)","Abdomen (rectum)","Phasic (distension)","36","Expectancy; saline infusion","Kotsis/Elsenbruch 2012 - confirm dataset.")

placebo("Freeman_2015_placebo","Freeman_2015_placebo",
"Distinct neural representations of placebo and nocebo effects","NeuroImage 2015;112:197-207. doi:10.1016/j.neuroimage.2015.03.015",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Forearm","Tonic","24","Conditioned; inert cream")

placebo("Geuter_2013_placebo","Geuter_2013_placebo",
"Cortical and subcortical responses to high and low effective placebo treatments","NeuroImage 2013;67:227-36. doi:10.1016/j.neuroimage.2012.11.029",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Forearm","Tonic","40","Conditioned; high/low inert cream")

placebo("Kong_2006_placebo","Kong_2006_placebo",
"Brain activity associated with expectancy-enhanced placebo analgesia as measured by functional magnetic resonance imaging","J Neurosci 2006;26(2):381-8. doi:10.1523/JNEUROSCI.3556-05.2006",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Forearm","Tonic","10","Conditioned; sham acupuncture")

placebo("Kong_2009_placebo","Kong_2009_placebo",
"Expectancy and treatment interactions: a dissociation between acupuncture analgesia and expectancy evoked placebo analgesia","NeuroImage 2009;45(3):940-9. doi:10.1016/j.neuroimage.2008.12.025",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Forearm","Tonic","12","Conditioned; sham acupuncture")

placebo("Lui_2010_placebo","Lui_2010_placebo",
"Neural bases of conditioned placebo analgesia","Pain 2010;151(3):816-824. doi:10.1016/j.pain.2010.09.021",
"Placebo+C|Placebo + conditioning","Laser","Hand/arm","Phasic (laser)","31","Conditioned; sham TENS")

placebo("Schenk_2014_placebo","Schenk_2014_placebo",
"Expectation requires treatment to boost pain relief: an fMRI study","Pain 2014;155(1):150-7. doi:10.1016/j.pain.2013.09.024",
"Placebo|Placebo (expectancy)","Heat/Capsaicin","Forearm","Tonic","32","Expectancy; lidocaine/inert cream")

placebo("Theysohn_2009_placebo","Theysohn_2014_placebo",
"Visceral placebo analgesia study (Essen group, Theysohn et al.)","Neurogastroenterol Motil ~2014 (verify)",
"Placebo|Placebo (expectancy)","Visceral (rectal distension)","Abdomen (rectum)","Phasic (distension)","30","Expectancy; saline infusion","Exact citation - verify (visceral placebo dataset).")

placebo("Wager_2004a_placebo","Wager_2004_study1_placebo",
"Placebo-induced changes in fMRI in the anticipation and experience of pain (Study 1)","Science 2004;303(5661):1162-7. doi:10.1126/science.1093065",
"Placebo|Placebo (expectancy)","Electrical (shock)","Hand/arm","Phasic","24","Expectancy; inert cream","Study 1 of Wager 2004.")

placebo("Wager_2004b_placebo","Wager_2004_study2_placebo",
"Placebo-induced changes in fMRI in the anticipation and experience of pain (Study 2)","Science 2004;303(5661):1162-7. doi:10.1126/science.1093065",
"Placebo|Placebo (expectancy)","Heat (thermal)","Forearm","Tonic","23","Expectancy/Conditioned; inert cream","Study 2 of Wager 2004. Zunhammer codes this as conditioned (user labeled 'Placebo').")

placebo("Wrobel_2014_placebo","Wrobel_2014_placebo",
"Haloperidol blocks dorsal striatum activity but not analgesia in a placebo paradigm","Cortex 2014;57:60-73. doi:10.1016/j.cortex.2014.02.023",
"Placebo+C|Placebo + conditioning","Heat (thermal)","Forearm","Tonic","38","Conditioned; inert cream")

wb = Workbook()
ws = wb.active
ws.title = "Studies"

hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
cell_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
grp_fills = {
 "Placebo + conditioning":"FCE4D6","Placebo (expectancy)":"FFF2CC",
 "Opioid (pharmacological)":"DDEBF7","Opioid (pharmacological) x expectancy":"DDEBF7",
 "Expectancy / Conditioning":"E2EFDA","Perceived control":"D9E1F2",
 "Cognitive self-regulation":"EDEDED","Cognitive self-regulation (uncertain)":"EDEDED",
 "Cognitive-affective (reward)":"EDEDED","Cognitive-affective (emotion)":"EDEDED",
 "Cognitive-affective (meaning) / Social":"EDEDED","Cognitive (environmental/attention)":"EDEDED",
 "Social":"FBE5D6","Social (social touch)":"FBE5D6","Mindfulness":"E1D5E7",
 "Placebo (+ conditioning)":"FCE4D6",
}

for j,h in enumerate(HEAD,1):
    c = ws.cell(1,j,h); c.fill=hdr_fill; c.font=hdr_font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border

for i,row in enumerate(R,2):
    for j,val in enumerate(row,1):
        c = ws.cell(i,j,val); c.font=cell_font; c.border=border
        c.alignment=Alignment(vertical="top",wrap_text=True)
    g = row[7]
    if g in grp_fills:
        ws.cell(i,8).fill = PatternFill("solid", fgColor=grp_fills[g])

widths=[10,22,30,34,40,34,18,24,30,18,18,16,16,18,22,22,12,40]
for j,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes="C2"
ws.row_dimensions[1].height=42
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{len(R)+1}"

lg = wb.create_sheet("Legend")
legend=[
 ["Study_summary - completed",""],
 ["Each paper looked up individually. Placebo rows (Bingel2006..Wrobel2014) sourced from Zunhammer et al. 2018, JAMA Neurol (NPS placebo meta-analysis).",""],
 ["",""],
 ["COLUMN","MEANING"],
 ["New Label","Lastname(first author)_year_contrast"],
 ["Contrast","Specific fMRI contrast yielding the analgesic effect"],
 ["Type of Intervention (validated, broad)","Validated broad mechanistic group (see groups below)"],
 ["Mechanism / subtype","More specific mechanism / manipulation"],
 ["Pain Modality","Heat / Electrical / Laser / Mechanical(pressure) / Visceral(rectal) / Chemical(capsaicin)"],
 ["Pain Temporality","Phasic (brief/transient) vs Tonic (sustained/state)"],
 ["N (analyzed)","Subjects in analyzed fMRI sample ('verify' = full text not accessible)"],
 ["Placebo Conditioning","Conditioned vs expectancy-only (placebo rows)"],
 ["",""],
 ["BROAD INTERVENTION GROUPS (validated)",""],
 ["Opioid (pharmacological)","Remifentanil / opioid drug"],
 ["Placebo (expectancy)","Placebo by verbal/expectancy only"],
 ["Placebo + conditioning","Placebo paired with conditioning"],
 ["Expectancy / Conditioning","Cue-based expectancy & classical/conceptual conditioning"],
 ["Perceived control","Controllability of the stimulus"],
 ["Cognitive self-regulation","Reappraisal / imagery / self-regulation"],
 ["Cognitive-affective","Reward, emotion, meaning manipulations"],
 ["Social","Social touch / social influence / prosocial meaning"],
 ["Mindfulness","Meditation / mindful acceptance"],
 ["",""],
 ["FLAGGED FOR VERIFICATION (full text not freely accessible)",""],
 ["Atlas_2012_remifentanil","N (JNeurosci paywalled)"],
 ["Brascher_2016_control","N (JNeurosci paywalled)"],
 ["Jepma_2015_conceptualConditioning","N (Psych Sci paywalled)"],
 ["Woo_2017 (ilcp_exp/ctrl)","Subset N of SIIPS-1 multi-study dataset"],
 ["LopezSola_2018_prosocialMeaning","N (Psychosom Med paywalled)"],
 ["Kim_2024_cueIntegration","Final analyzed N"],
 ["Riegner_2023_mindfulness","N (Pain paywalled)"],
 ["MPA2_heat / MPA2_pressure","Unpublished dataset - all fields"],
 ["Choi_2011 / Theysohn / Elsenbruch_2012","Exact original citation"],
]
for i,row in enumerate(legend,1):
    for j,val in enumerate(row,1):
        c=lg.cell(i,j,val); c.font=Font(name="Arial",size=10)
        c.alignment=Alignment(vertical="top",wrap_text=True)
lg.column_dimensions["A"].width=44; lg.column_dimensions["B"].width=72
lg.cell(1,1).font=Font(name="Arial",size=13,bold=True)
for hdr_row in (4,14,25):
    lg.cell(hdr_row,1).font=Font(name="Arial",size=10,bold=True)

wb.save("Study_summary_completed.xlsx")
print("Saved. Data rows:",len(R))
