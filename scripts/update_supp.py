# -*- coding: utf-8 -*-
# update_supp.py - fill rows now resolvable from supplements / Study-6 identity
import io, sys
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
FN = "Study_summary_completed.xlsx"
GREEN  = PatternFill("solid", fgColor="E2EFDA")
ORANGE = PatternFill("solid", fgColor="FCE4D6")

# cols 3..15: Site, Scanner, Field, Sequence, TR, TE, Flip, Voxel, Slices, FOV, Smoothing, Status, Note
SCAN = {
"Bingel_2011_remifentanil_posExpectancy": ["Hamburg / Oxford (Bingel/Tracey)","Varian / Siemens 3-T system (4-ch coil)","3 T",
  "T2* echo-planar (EPI)","3000","30","(n/r)","3 x 3 x 3 mm","42 (transverse)","224 x 224 mm; 64x64","8 mm","Verified",
  "Sci Transl Med 2011. From Suppl Methods. Slices tilted 30 deg. T1 MPRAGE 1 mm. 4-channel head coil."],

"Bingel_2011_placebo": ["Hamburg / Oxford","Varian / Siemens 3-T system (4-ch coil)","3 T",
  "T2* echo-planar (EPI)","3000","30","(n/r)","3 x 3 x 3 mm","42 (transverse)","224 x 224 mm; 64x64","8 mm","Verified",
  "Sci Transl Med 2011 (same study as Bingel_2011 remifentanil row). From Suppl Methods. Slice tilt 30 deg; T1 MPRAGE 1 mm."],

"Ellingsen_2013_placebo": ["Oslo / Gothenburg","Philips Achieva (8-ch SENSE R=2)","3 T",
  "Gradient-echo EPI (SENSE x2)","2000","30","80","3 x 3 mm in-plane; 0.3 mm gap","34 (axial)","240 x 240 mm","5 mm","Verified",
  "PNAS 2013. From SI. 510 volumes. T1: TR7.1/TE3.2/flip8/1x1 mm/160 slices."],

"Wager_2004_study1_placebo": ["Michigan (Wager)","Siemens 3.0 T head-dedicated","3 T",
  "T2* echo-planar (EPI)","3000","22","(n/r)","3.0 x 3.0 x 2.5 mm (1.5 mm gap)","30 (axial)","192 mm; 64x64","6 mm","Verified",
  "Science 2004 Study 1 (electrical shock). From SOM. 5 runs x 153 scans. Structural 256x256, FOV256, 128 x 1.33 mm sagittal."],

"Wager_2004_study2_placebo": ["Princeton (Wager)","GE Signa 3 T","3 T",
  "Spiral GRE (gradient-echo)","1500","20","90","3.75 x 3.75 x 5 mm (skip 0)","(n/r)","240 mm; 64x64","9 mm","Verified",
  "Science 2004 Study 2 (heat). From SOM. Spiral imaging (more OFC dropout than Study 1 EPI)."],

# ILCP = SIIPS-1 Study 6 (single novel dataset; both contrasts share acquisition)
"Woo_2017_expectancy / _perceivedControl": ["CANlab (Wager) - SIIPS-1 Study 6","(in Suppl. Table 1)","(in Suppl. Table 1)",
  "(in Suppl. Table 1)","(in Suppl. 1)","(in Suppl. 1)","(in Suppl. 1)","resampled to 2 x 2 x 2 mm","(in Suppl. 1)","(in Suppl. 1)","8 mm","Verified (partial)",
  "SIIPS-1 Study 6 (N=29): novel 2x2 perceived-control x expectancy dataset (left forearm heat). BOTH ILCP rows (expectancy + perceivedControl) come from this single dataset. Per-study scanner/TR/TE in Woo 2017 Supplementary Table 1 (not in main PDF). Common preprocessing: resample 2 mm, smooth 8 mm FWHM."],
}

# Studies sheet: update Woo 2017 N + notes (col 13 = N, col 18 = Notes)
STUDIES_EDITS = {
"Woo_2017_expectancy": {13: "29 (SIIPS-1 Study 6)",
  18: "ILCP = SIIPS-1 Study 6 (N=29), novel 2x2 control x expectancy dataset. Expectancy + perceivedControl contrasts share one dataset/acquisition."},
"Woo_2017_perceivedControl": {13: "29 (SIIPS-1 Study 6)",
  18: "ILCP = SIIPS-1 Study 6 (N=29), novel 2x2 control x expectancy dataset. Same dataset/acquisition as Woo_2017_expectancy."},
}

wb = openpyxl.load_workbook(FN)
ws = wb["Scanning parameters"]
labels = {ws.cell(r,1).value: r for r in range(2, ws.max_row+1)}
for label, vals in SCAN.items():
    r = labels.get(label)
    if not r:
        print("  !! not found in Scanning:", label); continue
    for j, v in enumerate(vals, start=3):
        ws.cell(r, j, v)
    ws.cell(r,14).fill = GREEN if vals[-2]=="Verified" else ORANGE
    print("  scan ok:", label)

ws2 = wb["Studies"]
labels2 = {ws2.cell(r,3).value: r for r in range(2, ws2.max_row+1)}
for label, edits in STUDIES_EDITS.items():
    r = labels2.get(label)
    if not r:
        print("  !! not found in Studies:", label); continue
    for col, val in edits.items():
        ws2.cell(r, col, val)
    print("  study ok:", label)

# refresh README counts
if "README" in wb.sheetnames:
    rm = wb["README"]
    ver = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,14).value=="Verified")
    part = sum(1 for r in range(2, ws.max_row+1) if str(ws.cell(r,14).value).startswith(("Verified (partial)","NOT RETRIEVED")))
    for r in range(1, rm.max_row+1):
        a = rm.cell(r,1).value
        if a=="Verified papers": rm.cell(r,2,str(ver))
        elif a and a.startswith("Not retrieved"): rm.cell(r,2,str(part))

wb.save(FN)
print("Saved", FN)
cnt={}
for r in range(2, ws.max_row+1):
    s=ws.cell(r,14).value; cnt[s]=cnt.get(s,0)+1
print("\nStatus counts:")
for k,v in sorted(cnt.items(), key=lambda x:-x[1]): print(f"  {v:3}  {k}")
